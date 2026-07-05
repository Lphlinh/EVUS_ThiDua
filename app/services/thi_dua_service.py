"""Core business service for monthly teacher scoring forms."""

from __future__ import annotations

from typing import Any

from gspread.exceptions import WorksheetNotFound

from app.models.thi_dua_detail import ThiDuaDetail
from app.models.thi_dua_header import ThiDuaHeader
from app.services.audit_service import (
    ACTION_SUBMIT_PHIEU,
    ACTION_BGH_UPDATE_SCORE,
    ACTION_BGH_FINALIZE_PHIEU,
    ACTION_BGH_SUMMARIZE_MONTH,
    write_audit_log,
)
from app.services.criteria_service import get_all_criteria, get_item_criteria
from app.services.google_sheets_service import (
    get_worksheet,
    open_spreadsheet,
    read_sheet_records,
    clear_sheet_records_cache,
)
from app.utils.date_utils import (
    get_previous_month,
    get_school_year_from_month,
    is_scoring_window_open,
    now_iso,
)

SHEET_TH = "TH_ThiDua"
SHEET_CT = "CT_ThiDua"
SHEET_TC = "DM_TieuChi"
SHEET_DM_GV = "DM_GiaoVien"
SHEET_TONGHOP = "TongHop_ThiDua"

STATUS_DANG_CHAM = 2
STATUS_DA_NOP = 3
STATUS_BGH_DA_CHINH_SUA = 4
STATUS_DA_CHOT = 5

TH_HEADERS = [
    "ID",
    "NamHoc",
    "Thang",
    "MaGV",
    "TrangThai",
    "TongDiem",
    "KhoaGV",
    "KhoaBGH",
    "NgayTao",
    "NgayCapNhat",
    "NgayNop",
    "NguoiCapNhatCuoi",
    "NguoiNop",
    "GhiChuBGH",
    "LanNop",
    "LanMoKhoa",
]

CT_HEADERS = [
    "ID",
    "IDPhieu",
    "MaTC",
    "DiemMacDinh",
    "DiemGV",
    "DiemBGH",
    "GhiChuGV",
    "GhiChuBGH",
    "NgayCapNhat",
    "NguoiCapNhat",
]

TONGHOP_HEADERS = [
    "ID",
    "NamHoc",
    "Thang",
    "MaGV",
    "HoTen",
    "ToChuyenMon",
    "TongDiem",
    "XepLoai",
    "NgayTongHop",
]


def get_current_scoring_month() -> str:
    """Return default scoring month in MM/YYYY format."""
    return get_previous_month()


def build_phieu_id(nam_hoc: str, thang: str, ma_gv: str) -> str:
    """Build stable primary key for TH_ThiDua."""
    safe_thang = str(thang).replace("/", "")
    return f"{nam_hoc}_{safe_thang}_{str(ma_gv).strip()}"




def _expected_phieu_id_from_values(nam_hoc: str, thang: str, ma_gv: str) -> str:
    """Build the only valid TH_ThiDua.ID for a teacher/month when data is sufficient."""
    nam_hoc = str(nam_hoc or "").strip()
    normalized_month = _normalize_month_key(thang)
    ma_gv = str(ma_gv or "").strip()
    if not nam_hoc or not normalized_month or not ma_gv:
        return ""
    return build_phieu_id(nam_hoc, normalized_month, ma_gv)


def _is_valid_phieu_id(row_id: str, nam_hoc: str, thang: str, ma_gv: str) -> bool:
    """Return True only when TH_ThiDua.ID matches the canonical business key."""
    expected_id = _expected_phieu_id_from_values(nam_hoc, thang, ma_gv)
    return bool(expected_id and str(row_id or "").strip() == expected_id)


def repair_th_ids_for_month(thang: str) -> dict:
    """Repair invalid TH_ThiDua.ID values for one month using NamHoc/Thang/MaGV.

    This is a guarded data repair, not a display patch.  It only rewrites ID
    when NamHoc, Thang and MaGV are all present and determine exactly one
    canonical key.  If rewriting would collide with an existing valid ID, the
    function stops and reports a clear error instead of guessing.
    """
    normalized_month = _normalize_month_key(thang)
    if not normalized_month:
        raise ValueError("Thiếu tháng cần kiểm tra khóa phiếu.")

    worksheet = get_worksheet(SHEET_TH)
    records = read_sheet_records(SHEET_TH)
    header_map = _get_header_map(worksheet)
    required_headers = ["ID", "NamHoc", "Thang", "MaGV"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ValueError("TH_ThiDua thiếu cột bắt buộc để kiểm tra ID: " + ", ".join(missing_headers))

    valid_id_owner: dict[str, int] = {}
    candidates: list[tuple[int, str, str, str]] = []

    for index, record in enumerate(records):
        row_number = index + 2
        canonical = _canonicalize_th_record(record)
        row_month = _normalize_month_key(canonical.get("Thang", ""))
        if row_month != normalized_month:
            continue

        nam_hoc = str(canonical.get("NamHoc", "")).strip()
        ma_gv = str(canonical.get("MaGV", "")).strip()
        current_id = str(canonical.get("ID", "")).strip()
        expected_id = _expected_phieu_id_from_values(nam_hoc, normalized_month, ma_gv)
        if not expected_id:
            candidates.append((row_number, current_id, expected_id, "Thiếu NamHoc/Thang/MaGV để dựng ID"))
            continue

        if current_id == expected_id:
            valid_id_owner[expected_id] = row_number
        else:
            candidates.append((row_number, current_id, expected_id, "Sai ID"))

    conflicts: list[str] = []
    pending_updates: list[dict[str, Any]] = []
    repaired: list[dict[str, Any]] = []
    id_column = header_map["ID"]

    for row_number, current_id, expected_id, reason in candidates:
        if not expected_id:
            conflicts.append(f"Dòng {row_number}: {reason}; ID hiện tại='{current_id}'.")
            continue
        owner = valid_id_owner.get(expected_id)
        if owner and owner != row_number:
            conflicts.append(
                f"Dòng {row_number}: ID hiện tại='{current_id}' cần sửa thành '{expected_id}' "
                f"nhưng ID này đã có ở dòng {owner}."
            )
            continue
        pending_updates.append({"range": _a1(row_number, id_column), "values": [[expected_id]]})
        repaired.append({"row": row_number, "old_id": current_id, "new_id": expected_id})
        valid_id_owner[expected_id] = row_number

    if conflicts:
        raise ValueError("Không thể tự sửa TH_ThiDua.ID vì có dữ liệu không nhất quán: " + " | ".join(conflicts))

    if pending_updates:
        worksheet.batch_update(pending_updates, value_input_option="RAW")
        clear_sheet_records_cache(SHEET_TH)

    return {
        "thang": normalized_month,
        "repaired_count": len(repaired),
        "repaired": repaired,
    }


def get_phieu(ma_gv: str, thang: str) -> dict | None:
    """Find one monthly scoring form by teacher and month.

    TH_ThiDua may use Vietnamese/legacy headers such as ``Mã GV`` and
    ``Tháng``. Always canonicalize the row before comparing, and normalize
    month strings so values like ``06/2026`` and ``6/2026`` match.
    """
    target_ma_gv = str(ma_gv).strip()
    target_thang = _normalize_month_key(thang)

    for record in read_sheet_records(SHEET_TH):
        canonical = _canonicalize_th_record(record)
        record_ma_gv = str(canonical.get("MaGV", "")).strip()
        record_thang = _normalize_month_key(canonical.get("Thang", ""))
        if record_ma_gv == target_ma_gv and record_thang == target_thang:
            return canonical
    return None


def get_phieu_list_by_month(thang: str) -> list[dict]:
    """Return de-duplicated monthly TH_ThiDua headers for BGH listing.

    This function only reads Google Sheets data. During M02 testing, duplicate
    TH_ThiDua rows can exist with the same ID. For display, keep only one row
    per form ID, preferring the row with a meaningful total score, then newer
    submit/update timestamps.
    """
    target_thang = _normalize_month_key(thang)
    selected_by_id: dict[str, dict] = {}
    order: list[str] = []

    for record in read_sheet_records(SHEET_TH):
        canonical = _canonicalize_th_record(record)
        record_thang = _normalize_month_key(canonical.get("Thang", ""))
        if record_thang != target_thang:
            continue

        row_id = str(canonical.get("ID", "")).strip()
        expected_id = _expected_phieu_id_from_values(
            str(canonical.get("NamHoc", "")).strip(),
            str(canonical.get("Thang", "")).strip(),
            str(canonical.get("MaGV", "")).strip(),
        )
        if expected_id and row_id != expected_id:
            # Do not silently trust a corrupted TH_ThiDua.ID such as "S".
            # The source repair is handled by repair_th_ids_for_month(); this
            # fallback keeps downstream logic on the canonical key after cache refresh.
            row_id = expected_id
            canonical["ID"] = row_id
        elif not row_id:
            row_id = expected_id
            canonical["ID"] = row_id

        if row_id not in selected_by_id:
            selected_by_id[row_id] = canonical
            order.append(row_id)
            continue

        if _th_record_priority(canonical) >= _th_record_priority(selected_by_id[row_id]):
            selected_by_id[row_id] = canonical

    result = [selected_by_id[row_id] for row_id in order]
    return sorted(
        result,
        key=lambda row: (
            str(row.get("MaGV", "")).strip(),
            str(row.get("ID", "")).strip(),
        ),
    )


def _th_record_priority(record: dict) -> tuple[int, str, str]:
    """Priority for selecting one TH_ThiDua row among duplicate IDs."""
    total = _safe_float(record.get("TongDiem"))
    has_total = int(total != 0)
    ngay_nop = str(record.get("NgayNop", "")).strip()
    ngay_cap_nhat = str(record.get("NgayCapNhat", "")).strip()
    return (has_total, ngay_nop, ngay_cap_nhat)


def create_phieu_for_teacher(ma_gv: str, thang: str | None = None) -> dict:
    """Create TH_ThiDua and CT_ThiDua for a teacher.

    This function must only be called after permission/window checks.
    """
    thang = thang or get_current_scoring_month()
    nam_hoc = get_school_year_from_month(thang)
    id_phieu = build_phieu_id(nam_hoc, thang, ma_gv)

    existing = get_phieu(ma_gv, thang)
    if existing:
        return existing

    timestamp = now_iso()
    header = ThiDuaHeader(
        ID=id_phieu,
        NamHoc=nam_hoc,
        Thang=thang,
        MaGV=str(ma_gv).strip(),
        TrangThai=STATUS_DANG_CHAM,
        TongDiem=0,
        KhoaGV=False,
        KhoaBGH=False,
        NgayTao=timestamp,
        NgayCapNhat=timestamp,
        NgayNop="",
        NguoiCapNhatCuoi=str(ma_gv).strip(),
        NguoiNop="",
        GhiChuBGH="",
        LanNop=0,
        LanMoKhoa=0,
    )

    get_worksheet(SHEET_TH).append_row(
        [header.to_row().get(column, "") for column in TH_HEADERS],
        value_input_option="USER_ENTERED",
    )
    generate_chi_tiet_from_tieu_chi(id_phieu, actor=ma_gv)
    return header.to_row()


def generate_chi_tiet_from_tieu_chi(id_phieu: str, actor: str = "SYSTEM") -> list[dict]:
    """Generate missing CT_ThiDua rows from normalized ITEM criteria.

    Safety rule: never append a detail row if the same CT_ThiDua.ID already
    exists. This prevents duplicate ITEM rows for the same monthly form.
    """
    id_phieu = str(id_phieu).strip()
    timestamp = now_iso()
    details: list[dict] = []

    existing_ids = {
        _get_record_text(record, "ID") or str(record.get("ID", "")).strip()
        for record in read_sheet_records(SHEET_CT)
        if _get_record_text(record, "IDPhieu") == id_phieu
    }

    for criterion in get_item_criteria():
        ma_tc = str(criterion.get("ma_tc", "")).strip()
        if not ma_tc:
            continue

        detail_id = f"{id_phieu}_{ma_tc}"
        if detail_id in existing_ids:
            continue

        diem_mac_dinh = float(criterion.get("diem_mac_dinh") or 0)
        detail = ThiDuaDetail(
            ID=detail_id,
            IDPhieu=id_phieu,
            MaTC=ma_tc,
            DiemMacDinh=diem_mac_dinh,
            DiemGV=diem_mac_dinh,
            DiemBGH="",
            GhiChuGV="",
            GhiChuBGH="",
            NgayCapNhat=timestamp,
            NguoiCapNhat=str(actor).strip(),
        )
        details.append(detail.to_row())
        existing_ids.add(detail_id)

    if details:
        get_worksheet(SHEET_CT).append_rows(
            [[row.get(column, "") for column in CT_HEADERS] for row in details],
            value_input_option="USER_ENTERED",
        )

    return details


def sync_chi_tiet_from_tieu_chi(id_phieu: str, actor: str = "SYSTEM") -> None:
    """Synchronize only missing/default fields without overwriting DiemGV.

    This repair function must not create duplicate CT_ThiDua rows and must not
    reset teacher-entered scores. It appends only truly missing ITEM IDs.
    """
    id_phieu = str(id_phieu).strip()
    if not id_phieu:
        return

    item_criteria = get_item_criteria()
    criteria_by_code = {str(item.get("ma_tc", "")).strip(): item for item in item_criteria}
    all_records = read_sheet_records(SHEET_CT)
    current_rows = [record for record in all_records if _get_record_text(record, "IDPhieu") == id_phieu]
    existing_ids = {_get_record_text(record, "ID") for record in current_rows if _get_record_text(record, "ID")}

    worksheet = get_worksheet(SHEET_CT)
    header_map = _get_header_map(worksheet)
    timestamp = now_iso()

    rows_to_append: list[dict] = []
    for ma_tc, criterion in criteria_by_code.items():
        detail_id = f"{id_phieu}_{ma_tc}"
        if detail_id in existing_ids:
            continue
        diem_mac_dinh = float(criterion.get("diem_mac_dinh") or 0)
        detail = ThiDuaDetail(
            ID=detail_id,
            IDPhieu=id_phieu,
            MaTC=ma_tc,
            DiemMacDinh=diem_mac_dinh,
            DiemGV=diem_mac_dinh,
            DiemBGH="",
            GhiChuGV="",
            GhiChuBGH="",
            NgayCapNhat=timestamp,
            NguoiCapNhat=str(actor).strip(),
        )
        rows_to_append.append(detail.to_row())
        existing_ids.add(detail_id)

    if rows_to_append:
        worksheet.append_rows(
            [[row.get(column, "") for column in CT_HEADERS] for row in rows_to_append],
            value_input_option="USER_ENTERED",
        )

    refreshed_records = read_sheet_records(SHEET_CT)
    row_map: dict[str, int] = {}
    for index, record in enumerate(refreshed_records):
        record_id = _get_record_text(record, "ID")
        if record_id and record_id not in row_map:
            row_map[record_id] = index + 2

    for record in refreshed_records:
        if _get_record_text(record, "IDPhieu") != id_phieu:
            continue

        ma_tc = _get_record_text(record, "MaTC")
        criterion = criteria_by_code.get(ma_tc)
        if not criterion:
            continue

        expected_default = float(criterion.get("diem_mac_dinh") or 0)
        current_default = _safe_float(_get_record_text(record, "DiemMacDinh"))
        row_number = row_map.get(_get_record_text(record, "ID"))
        if not row_number:
            continue

        updates: dict[str, Any] = {}
        if current_default != expected_default:
            updates["DiemMacDinh"] = expected_default

        if updates:
            updates["NgayCapNhat"] = timestamp
            updates["NguoiCapNhat"] = actor
            _update_cells_by_headers(worksheet, row_number, header_map, updates)


def get_chi_tiet_phieu(id_phieu: str) -> list[dict]:
    """Return de-duplicated detail rows for one form without modifying CT_ThiDua."""
    records = [
        record
        for record in read_sheet_records(SHEET_CT)
        if _get_record_text(record, "IDPhieu") == str(id_phieu).strip()
    ]
    return _dedupe_ct_records(records)



def _ct_record_priority(record: dict) -> tuple[int, str]:
    canonical = _canonicalize_ct_record(record)
    diem_mac_dinh = _safe_float(canonical.get("DiemMacDinh"))
    diem_gv = _safe_float(canonical.get("DiemGV"))
    has_teacher_change = int(
        str(canonical.get("GhiChuGV", "")).strip() != ""
        or diem_gv != diem_mac_dinh
    )
    return (has_teacher_change, str(canonical.get("NgayCapNhat", "")).strip())


def _dedupe_ct_records(records: list[dict]) -> list[dict]:
    """Keep one row per CT_ThiDua.ID, preferring teacher-edited data."""
    selected: dict[str, dict] = {}
    order: list[str] = []
    for record in records:
        canonical = _canonicalize_ct_record(record)
        row_id = str(canonical.get("ID") or f"{canonical.get('IDPhieu', '')}_{canonical.get('MaTC', '')}").strip()
        if not row_id:
            continue
        if row_id not in selected:
            selected[row_id] = record
            order.append(row_id)
            continue
        if _ct_record_priority(record) >= _ct_record_priority(selected[row_id]):
            selected[row_id] = record
    return [selected[row_id] for row_id in order]

def get_readonly_phieu(id_phieu: str) -> dict:
    """Return one scoring form for read-only display.

    The returned object contains the canonical TH_ThiDua header and the
    de-duplicated CT_ThiDua details. This function does not write or repair
    data in Google Sheets.
    """
    id_phieu = str(id_phieu).strip()
    if not id_phieu:
        raise ValueError("Thiếu ID phiếu thi đua cần xem.")

    header = get_header_by_id(id_phieu)
    if not header:
        raise ValueError(f"Không tìm thấy TH_ThiDua.ID = {id_phieu}")

    details = get_chi_tiet_phieu(id_phieu)
    return {
        "header": header,
        "details": details,
    }


def calculate_total_score(details: list[dict]) -> float:
    """Calculate total score using official score rule."""
    total = 0.0
    seen: set[str] = set()
    for detail in details:
        record = _canonicalize_ct_record(detail)
        row_id = str(record.get("ID") or f"{record.get('IDPhieu', '')}_{record.get('MaTC', '')}").strip()
        if row_id and row_id in seen:
            continue
        if row_id:
            seen.add(row_id)
        diem_bgh = str(record.get("DiemBGH", "")).strip()
        if diem_bgh != "":
            total += _safe_float(diem_bgh)
        else:
            total += _safe_float(record.get("DiemGV"))
    return total



def save_bgh_scores(id_phieu: str, changed_rows: list[dict], ma_bgh: str) -> float:
    """Save BGH score adjustments into CT_ThiDua and update TH_ThiDua.

    BGH is only allowed to update DiemBGH and GhiChuBGH. If DiemBGH equals
    DiemGV, DiemBGH is stored as blank so the official score still follows
    the agreed rule: use DiemBGH when present, otherwise use DiemGV.
    """
    id_phieu = str(id_phieu).strip()
    ma_bgh = str(ma_bgh).strip()
    if not id_phieu:
        raise ValueError("Thiếu ID phiếu thi đua cần lưu điểm BGH.")
    if not changed_rows:
        return calculate_total_score(get_chi_tiet_phieu(id_phieu))

    phieu = get_header_by_id(id_phieu)
    if not phieu:
        raise ValueError(f"Không tìm thấy TH_ThiDua.ID = {id_phieu}")
    if _safe_int(phieu.get("TrangThai"), default=0) == STATUS_DA_CHOT or _parse_bool(phieu.get("KhoaBGH")):
        raise ValueError("Phiếu đã chốt hoặc đã khóa BGH, không thể chỉnh điểm.")

    worksheet = get_worksheet(SHEET_CT)
    all_records = read_sheet_records(SHEET_CT)
    header_map = _get_canonical_header_map(worksheet)
    timestamp = now_iso()

    required_headers = ["ID", "IDPhieu", "MaTC", "DiemGV", "DiemBGH", "GhiChuBGH"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ValueError("CT_ThiDua thiếu cột bắt buộc để lưu điểm BGH: " + ", ".join(missing_headers))

    row_map_by_id: dict[str, int] = {}
    row_map_by_key: dict[tuple[str, str], int] = {}
    for index, record in enumerate(all_records):
        row_number = index + 2
        record_id = _get_record_text(record, "ID")
        record_id_phieu = _get_record_text(record, "IDPhieu")
        record_ma_tc = _get_record_text(record, "MaTC")
        if record_id:
            row_map_by_id[record_id] = row_number
        if record_id_phieu and record_ma_tc:
            row_map_by_key[(record_id_phieu, record_ma_tc)] = row_number

    current_records = _dedupe_ct_records([
        record for record in all_records if _get_record_text(record, "IDPhieu") == id_phieu
    ])
    working_by_ma_tc: dict[str, dict] = {}
    for record in current_records:
        canonical = _canonicalize_ct_record(record)
        ma_tc = str(canonical.get("MaTC", "")).strip()
        if ma_tc:
            working_by_ma_tc[ma_tc] = canonical

    pending_updates: list[dict[str, Any]] = []
    changed_count = 0

    for changed in changed_rows:
        ma_tc = str(changed.get("MaTC", "")).strip()
        if not ma_tc:
            continue
        row_id = str(changed.get("ID", "")).strip()
        row_number = row_map_by_id.get(row_id) or row_map_by_key.get((id_phieu, ma_tc))
        if not row_number:
            raise ValueError(f"Không tìm thấy dòng CT_ThiDua để lưu điểm BGH: ID={row_id}, MaTC={ma_tc}")

        current = dict(working_by_ma_tc.get(ma_tc, {}))
        diem_gv = _safe_float(current.get("DiemGV"))
        input_bgh = _safe_float(changed.get("DiemBGH"))
        ghi_chu_bgh = str(changed.get("GhiChuBGH", "")).strip()
        diem_bgh_to_store: Any = "" if input_bgh == diem_gv else input_bgh

        updates = {
            "DiemBGH": diem_bgh_to_store,
            "GhiChuBGH": ghi_chu_bgh,
            "NgayCapNhat": timestamp,
            "NguoiCapNhat": ma_bgh,
        }
        for canonical_name, value in updates.items():
            column_number = header_map.get(canonical_name)
            if column_number:
                pending_updates.append({"range": _a1(row_number, column_number), "values": [[value]]})

        current.update({
            "DiemBGH": diem_bgh_to_store,
            "GhiChuBGH": ghi_chu_bgh,
            "NgayCapNhat": timestamp,
            "NguoiCapNhat": ma_bgh,
        })
        working_by_ma_tc[ma_tc] = current
        changed_count += 1

    if changed_count == 0:
        return calculate_total_score(list(working_by_ma_tc.values()))

    if pending_updates:
        worksheet.batch_update(pending_updates, value_input_option="USER_ENTERED")
        clear_sheet_records_cache(SHEET_CT)

    total = calculate_total_score(list(working_by_ma_tc.values()))
    update_header_fields(
        id_phieu,
        {
            "TrangThai": STATUS_BGH_DA_CHINH_SUA,
            "TongDiem": total,
            "NgayCapNhat": timestamp,
            "NguoiCapNhatCuoi": ma_bgh,
        },
    )
    clear_sheet_records_cache(SHEET_TH)
    clear_sheet_records_cache(SHEET_CT)
    write_audit_log(
        ACTION_BGH_UPDATE_SCORE,
        id_phieu,
        ma_bgh,
        f"BGH chỉnh điểm thi đua. Số dòng cập nhật: {changed_count}. Tổng điểm: {total}",
    )
    return total


def finalize_month(thang: str, ma_bgh: str) -> dict:
    """Finalize all submitted/BGH-edited forms for one month.

    Only forms with TrangThai 3 or 4 are finalized. Draft/unsubmitted forms
    are left unchanged so BGH can see which teachers have not submitted.
    """
    normalized_month = _normalize_month_key(thang)
    ma_bgh = str(ma_bgh).strip() or "BGH"
    if not normalized_month:
        raise ValueError("Thiếu tháng cần chốt.")

    phieu_list = get_phieu_list_by_month(normalized_month)
    eligible = [
        phieu for phieu in phieu_list
        if _safe_int(phieu.get("TrangThai"), default=0) in {STATUS_DA_NOP, STATUS_BGH_DA_CHINH_SUA}
    ]
    already_finalized = [
        phieu for phieu in phieu_list
        if _safe_int(phieu.get("TrangThai"), default=0) == STATUS_DA_CHOT
    ]
    unsubmitted = [
        phieu for phieu in phieu_list
        if _safe_int(phieu.get("TrangThai"), default=0) not in {
            STATUS_DA_NOP,
            STATUS_BGH_DA_CHINH_SUA,
            STATUS_DA_CHOT,
        }
    ]

    if not eligible:
        return {
            "thang": normalized_month,
            "finalized_count": 0,
            "already_finalized_count": len(already_finalized),
            "unsubmitted_count": len(unsubmitted),
            "finalized_ids": [],
        }

    timestamp = now_iso()
    finalized_ids: list[str] = []
    for phieu in eligible:
        id_phieu = str(phieu.get("ID", "")).strip()
        if not id_phieu:
            continue
        update_header_fields(
            id_phieu,
            {
                "TrangThai": STATUS_DA_CHOT,
                "KhoaGV": True,
                "KhoaBGH": True,
                "NgayCapNhat": timestamp,
                "NguoiCapNhatCuoi": ma_bgh,
            },
        )
        finalized_ids.append(id_phieu)

    clear_sheet_records_cache(SHEET_TH)
    write_audit_log(
        ACTION_BGH_FINALIZE_PHIEU,
        normalized_month,
        ma_bgh,
        (
            f"BGH chốt tháng {normalized_month}. "
            f"Số phiếu chốt: {len(finalized_ids)}. "
            f"Số phiếu chưa nộp/bỏ qua: {len(unsubmitted)}."
        ),
    )
    clear_sheet_records_cache(SHEET_TH)

    return {
        "thang": normalized_month,
        "finalized_count": len(finalized_ids),
        "already_finalized_count": len(already_finalized),
        "unsubmitted_count": len(unsubmitted),
        "finalized_ids": finalized_ids,
    }


def summarize_month(thang: str, ma_bgh: str) -> dict:
    """Create/update monthly summary rows in TongHop_ThiDua.

    One finalized form creates one summary row. Running this function again
    updates existing rows by ID and does not create duplicates.
    """
    normalized_month = _normalize_month_key(thang)
    ma_bgh = str(ma_bgh).strip() or "BGH"
    if not normalized_month:
        raise ValueError("Thiếu tháng cần tổng hợp.")

    phieu_list = get_phieu_list_by_month(normalized_month)
    finalized = [
        phieu for phieu in phieu_list
        if _safe_int(phieu.get("TrangThai"), default=0) == STATUS_DA_CHOT
    ]
    if not finalized:
        return {
            "thang": normalized_month,
            "summary_count": 0,
            "created_count": 0,
            "updated_count": 0,
        }

    teacher_map = _get_teacher_master_map()
    worksheet = _get_or_create_tonghop_worksheet()
    _ensure_tonghop_headers(worksheet)
    header_map = _get_tonghop_header_map(worksheet)

    required_headers = ["ID", "NamHoc", "Thang", "MaGV", "HoTen", "ToChuyenMon", "TongDiem", "XepLoai", "NgayTongHop"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ValueError("TongHop_ThiDua thiếu cột bắt buộc: " + ", ".join(missing_headers))

    existing_records = read_sheet_records(SHEET_TONGHOP)
    existing_row_by_id: dict[str, int] = {}
    existing_row_by_key: dict[tuple[str, str, str], int] = {}
    for index, record in enumerate(existing_records):
        row_number = index + 2
        row_id = _get_tonghop_record_text(record, "ID")
        if row_id and row_id not in existing_row_by_id:
            existing_row_by_id[row_id] = row_number

        nam_hoc = _get_tonghop_record_text(record, "NamHoc")
        row_thang = _normalize_month_key(_get_tonghop_record_text(record, "Thang"))
        ma_gv = _get_tonghop_record_text(record, "MaGV")
        row_key = (nam_hoc, row_thang, ma_gv)
        if nam_hoc and row_thang and ma_gv and row_key not in existing_row_by_key:
            existing_row_by_key[row_key] = row_number

    timestamp = now_iso()
    pending_updates: list[dict[str, Any]] = []
    rows_to_append: list[list[Any]] = []
    created_count = 0
    updated_count = 0

    for phieu in finalized:
        nam_hoc = str(phieu.get("NamHoc", "")).strip()
        ma_gv = str(phieu.get("MaGV", "")).strip()
        id_phieu = _build_safe_summary_id(phieu, normalized_month)
        if not id_phieu or not ma_gv:
            continue
        teacher = teacher_map.get(ma_gv, {})
        summary_row = {
            "ID": id_phieu,
            "NamHoc": nam_hoc,
            "Thang": normalized_month,
            "MaGV": ma_gv,
            "HoTen": str(teacher.get("HoTen", "")).strip(),
            "ToChuyenMon": str(teacher.get("ToChuyenMon", "")).strip(),
            "TongDiem": _safe_float(phieu.get("TongDiem")),
            "XepLoai": "",
            "NgayTongHop": timestamp,
        }

        row_key = (nam_hoc, normalized_month, ma_gv)
        row_number = existing_row_by_id.get(id_phieu) or existing_row_by_key.get(row_key)
        if row_number:
            for field, value in summary_row.items():
                column_number = header_map.get(field)
                if column_number:
                    pending_updates.append({"range": _a1(row_number, column_number), "values": [[value]]})
            updated_count += 1
        else:
            rows_to_append.append([summary_row.get(column, "") for column in TONGHOP_HEADERS])
            created_count += 1

    if pending_updates:
        worksheet.batch_update(pending_updates, value_input_option="RAW")
    if rows_to_append:
        worksheet.append_rows(rows_to_append, value_input_option="RAW")

    clear_sheet_records_cache(SHEET_TONGHOP)
    write_audit_log(
        ACTION_BGH_SUMMARIZE_MONTH,
        normalized_month,
        ma_bgh,
        (
            f"BGH tổng hợp tháng {normalized_month}. "
            f"Số dòng tổng hợp: {created_count + updated_count}. "
            f"Thêm mới: {created_count}. Cập nhật: {updated_count}."
        ),
    )

    return {
        "thang": normalized_month,
        "summary_count": created_count + updated_count,
        "created_count": created_count,
        "updated_count": updated_count,
    }




def get_month_summary_rows(thang: str) -> list[dict]:
    """Return TongHop_ThiDua rows for one month using canonical keys.

    This function only reads summary data. It does not create or modify any
    worksheet. Excel export must rely on TongHop_ThiDua and must not recalculate
    scores from CT_ThiDua.
    """
    normalized_month = _normalize_month_key(thang)
    if not normalized_month:
        raise ValueError("Thiếu tháng cần xuất Excel.")

    try:
        records = read_sheet_records(SHEET_TONGHOP)
    except WorksheetNotFound as exc:
        raise ValueError(
            "Chưa có dữ liệu tổng hợp tháng. Vui lòng thực hiện Tổng hợp tháng trước khi xuất Excel."
        ) from exc

    rows: list[dict] = []
    for record in records:
        row_thang = _normalize_month_key(_get_tonghop_record_text(record, "Thang"))
        if row_thang != normalized_month:
            continue
        rows.append({
            "ID": _get_tonghop_record_text(record, "ID"),
            "NamHoc": _get_tonghop_record_text(record, "NamHoc"),
            "Thang": row_thang,
            "MaGV": _get_tonghop_record_text(record, "MaGV"),
            "HoTen": _get_tonghop_record_text(record, "HoTen"),
            "ToChuyenMon": _get_tonghop_record_text(record, "ToChuyenMon"),
            "TongDiem": _safe_float(_get_tonghop_record_text(record, "TongDiem")),
            "XepLoai": _get_tonghop_record_text(record, "XepLoai"),
            "NgayTongHop": _get_tonghop_record_text(record, "NgayTongHop"),
        })

    rows.sort(key=lambda item: (str(item.get("ToChuyenMon", "")), str(item.get("MaGV", ""))))
    return rows


def build_monthly_excel_export(thang: str) -> tuple[str, bytes, int]:
    """Build monthly all-school Excel export from TongHop_ThiDua.

    Returns: (filename, bytes, row_count). The export is generated in memory and
    does not write to Google Sheets or local project files.
    """
    normalized_month = _normalize_month_key(thang)
    rows = get_month_summary_rows(normalized_month)
    if not rows:
        raise ValueError("Chưa có dữ liệu tổng hợp tháng. Vui lòng thực hiện Tổng hợp tháng trước khi xuất Excel.")

    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("Thiếu thư viện openpyxl. Vui lòng cài đặt dependency trước khi xuất Excel.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TongHop_Thang"

    title = f"BẢNG TỔNG HỢP THI ĐUA THÁNG {normalized_month}"
    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 24

    worksheet.merge_cells("A2:G2")
    worksheet["A2"] = f"Ngày xuất: {now_iso()}"
    worksheet["A2"].alignment = Alignment(horizontal="right")
    worksheet["A2"].font = Font(italic=True, size=10)

    headers = ["STT", "Mã GV", "Họ tên", "Tổ chuyên môn", "Tổng điểm", "Xếp loại", "Ghi chú"]
    header_row = 4
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_index, row in enumerate(rows, start=header_row + 1):
        excel_values = [
            row_index - header_row,
            row.get("MaGV", ""),
            row.get("HoTen", ""),
            row.get("ToChuyenMon", ""),
            row.get("TongDiem", 0),
            row.get("XepLoai", ""),
            "",
        ]
        for column_index, value in enumerate(excel_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if column_index in {1, 2, 5, 6}:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if column_index == 5:
                cell.number_format = "0.##"

    for column_index in range(1, len(headers) + 1):
        worksheet.cell(row=header_row, column=column_index).border = border

    widths = {
        1: 8,
        2: 14,
        3: 28,
        4: 24,
        5: 14,
        6: 14,
        7: 28,
    }
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:G{header_row + len(rows)}"

    output = BytesIO()
    workbook.save(output)
    safe_month = normalized_month.replace("/", "_")
    filename = f"TongHop_ThiDua_Thang_{safe_month}.xlsx"
    return filename, output.getvalue(), len(rows)

def _build_safe_summary_id(phieu: dict, normalized_month: str) -> str:
    """Return stable TongHop_ThiDua.ID from the TH_ThiDua business key.

    Some summary rows were previously written with an invalid short ID such as
    "S".  The monthly summary key must always follow TH_ThiDua.ID format:
    NamHoc_MMYYYY_MaGV.
    """
    raw_id = str(phieu.get("ID", "")).strip()
    nam_hoc = str(phieu.get("NamHoc", "")).strip()
    ma_gv = str(phieu.get("MaGV", "")).strip()
    expected_id = build_phieu_id(nam_hoc, normalized_month, ma_gv) if nam_hoc and ma_gv else ""

    if expected_id and (not raw_id or len(raw_id) < 10 or raw_id.upper() == "S"):
        return expected_id
    if expected_id and ma_gv and not raw_id.endswith(f"_{ma_gv}"):
        return expected_id
    return raw_id or expected_id

def _get_or_create_tonghop_worksheet() -> Any:
    """Return TongHop_ThiDua worksheet; create it with standard headers if missing."""
    try:
        return get_worksheet(SHEET_TONGHOP)
    except WorksheetNotFound:
        spreadsheet = open_spreadsheet()
        spreadsheet.add_worksheet(
            title=SHEET_TONGHOP,
            rows=1000,
            cols=len(TONGHOP_HEADERS),
        )
        worksheet = get_worksheet(SHEET_TONGHOP)
        _ensure_tonghop_headers(worksheet)
        clear_sheet_records_cache(SHEET_TONGHOP)
        return worksheet


def _ensure_tonghop_headers(worksheet: Any) -> None:
    headers = [str(header).strip() for header in worksheet.row_values(1)]
    if any(headers):
        return
    worksheet.update("A1:I1", [TONGHOP_HEADERS], value_input_option="RAW")
    clear_sheet_records_cache(SHEET_TONGHOP)


def _get_tonghop_header_map(worksheet: Any) -> dict[str, int]:
    headers = worksheet.row_values(1)
    normalized_headers = {_normalize_key(header): index + 1 for index, header in enumerate(headers) if str(header).strip()}
    aliases = {
        "ID": ("ID", "id"),
        "NamHoc": ("NamHoc", "Năm học", "nam_hoc", "namhoc"),
        "Thang": ("Thang", "Tháng", "thang"),
        "MaGV": ("MaGV", "Mã GV", "ma_gv", "magv"),
        "HoTen": ("HoTen", "Họ tên", "Họ và tên", "ho_ten", "hoten"),
        "ToChuyenMon": ("ToChuyenMon", "Tổ chuyên môn", "To chuyen mon", "to_chuyen_mon", "tochuyenmon"),
        "TongDiem": ("TongDiem", "Tổng điểm", "tong_diem", "tongdiem"),
        "XepLoai": ("XepLoai", "Xếp loại", "xep_loai", "xeploai"),
        "NgayTongHop": ("NgayTongHop", "Ngày tổng hợp", "ngay_tong_hop", "ngaytonghop"),
    }
    result: dict[str, int] = {}
    for canonical, names in aliases.items():
        for name in names:
            column_number = normalized_headers.get(_normalize_key(name))
            if column_number:
                result[canonical] = column_number
                break
    return result


def _get_tonghop_record_text(record: dict, canonical: str) -> str:
    aliases = {
        "ID": ("ID", "id"),
        "NamHoc": ("NamHoc", "Năm học", "nam_hoc", "namhoc"),
        "Thang": ("Thang", "Tháng", "thang"),
        "MaGV": ("MaGV", "Mã GV", "ma_gv", "magv"),
        "HoTen": ("HoTen", "Họ tên", "Họ và tên", "ho_ten", "hoten"),
        "ToChuyenMon": ("ToChuyenMon", "Tổ chuyên môn", "to_chuyen_mon", "tochuyenmon"),
        "TongDiem": ("TongDiem", "Tổng điểm", "tong_diem", "tongdiem"),
        "XepLoai": ("XepLoai", "Xếp loại", "xep_loai", "xeploai"),
        "NgayTongHop": ("NgayTongHop", "Ngày tổng hợp", "ngay_tong_hop", "ngaytonghop"),
    }.get(canonical, (canonical,))
    return _get_text(record, *aliases)


def _get_teacher_master_map() -> dict[str, dict]:
    teachers: dict[str, dict] = {}
    for record in read_sheet_records(SHEET_DM_GV):
        ma_gv = _get_text(record, "ma_gv", "MaGV", "Mã GV", "Mã định danh", "ma_dinh_danh")
        if not ma_gv:
            continue
        teachers[ma_gv] = {
            "MaGV": ma_gv,
            "HoTen": _get_text(record, "ho_ten", "HoTen", "Họ tên", "Họ và tên", "Họ tên Giáo viên"),
            "ToChuyenMon": _get_text(record, "to_chuyen_mon", "ToChuyenMon", "Tổ chuyên môn", "Tổ", "to"),
        }
    return teachers

def save_teacher_scores(id_phieu: str, changed_rows: list[dict], ma_gv: str) -> float:
    """Save teacher scores and update TH_ThiDua summary.

    This implementation writes only real CT_ThiDua rows for the current form.
    Header names are resolved by aliases so the code still works if the sheet
    uses legacy/variant labels for the same business columns.
    """
    id_phieu = str(id_phieu).strip()
    if not changed_rows:
        details = get_chi_tiet_phieu(id_phieu)
        return calculate_total_score(details)

    worksheet = get_worksheet(SHEET_CT)
    all_records = read_sheet_records(SHEET_CT)
    exact_header_map = _get_header_map(worksheet)
    header_map = _get_canonical_header_map(worksheet)
    timestamp = now_iso()

    # CT_ThiDua hiện tại có cấu trúc legacy từ Form v14.
    # Chỉ bắt buộc các cột thật sự cần để lưu điểm GV; cột metadata
    # NgayCapNhat/NguoiCapNhat nếu chưa có trong Google Sheets thì bỏ qua.
    required_headers = ["ID", "IDPhieu", "MaTC", "DiemGV", "GhiChuGV"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ValueError("CT_ThiDua thiếu cột bắt buộc: " + ", ".join(missing_headers))

    row_map_by_id: dict[str, int] = {}
    row_priority_by_id: dict[str, tuple[int, str]] = {}
    row_map_by_key: dict[tuple[str, str], int] = {}
    row_priority_by_key: dict[tuple[str, str], tuple[int, str]] = {}
    for index, record in enumerate(all_records):
        row_number = index + 2
        row_id_key = _get_record_text(record, "ID")
        row_key = (_get_record_text(record, "IDPhieu"), _get_record_text(record, "MaTC"))
        priority = _ct_record_priority(record)
        if row_id_key and (row_id_key not in row_map_by_id or priority >= row_priority_by_id[row_id_key]):
            row_map_by_id[row_id_key] = row_number
            row_priority_by_id[row_id_key] = priority
        if row_key[0] and row_key[1] and (row_key not in row_map_by_key or priority >= row_priority_by_key[row_key]):
            row_map_by_key[row_key] = row_number
            row_priority_by_key[row_key] = priority

    updated_count = 0
    pending_updates: list[dict[str, Any]] = []
    working_details = []
    current_by_key = {}
    current_records = _dedupe_ct_records([
        record for record in all_records if _get_record_text(record, "IDPhieu") == id_phieu
    ])
    for record in current_records:
        row = _canonicalize_ct_record(record)
        working_details.append(row)
        current_by_key[str(row.get("MaTC", "")).strip()] = row

    for changed in changed_rows:
        ma_tc = str(changed.get("MaTC", "")).strip()
        if not ma_tc:
            continue

        row_id = str(changed.get("ID", "")).strip()
        row_number = row_map_by_id.get(row_id) or row_map_by_key.get((id_phieu, ma_tc))
        if not row_number:
            continue

        updates = {
            "DiemGV": _safe_float(changed.get("DiemGV")),
            "GhiChuGV": str(changed.get("GhiChuGV", "")).strip(),
            "NgayCapNhat": timestamp,
            "NguoiCapNhat": ma_gv,
        }
        if "DiemMacDinh" in header_map:
            updates["DiemMacDinh"] = _safe_float(changed.get("DiemMacDinh"))

        # Ghi theo cột nghiệp vụ đã chuẩn hóa; không đụng DiemBGH/GhiChuBGH.
        for canonical_name, value in updates.items():
            column_number = header_map.get(canonical_name)
            if column_number:
                pending_updates.append({
                    "range": _a1(row_number, column_number),
                    "values": [[value]],
                })

        updated_count += 1
        current = dict(current_by_key.get(ma_tc, {}))
        current.update({
            "ID": row_id or current.get("ID", f"{id_phieu}_{ma_tc}"),
            "IDPhieu": id_phieu,
            "MaTC": ma_tc,
            "DiemMacDinh": _safe_float(changed.get("DiemMacDinh", current.get("DiemMacDinh", 0))),
            "DiemGV": _safe_float(changed.get("DiemGV")),
            "GhiChuGV": str(changed.get("GhiChuGV", "")).strip(),
            "DiemBGH": current.get("DiemBGH", ""),
            "GhiChuBGH": current.get("GhiChuBGH", ""),
            "NgayCapNhat": timestamp,
            "NguoiCapNhat": ma_gv,
        })
        current_by_key[ma_tc] = current

    if updated_count == 0:
        raise ValueError("Không có dòng CT_ThiDua nào được cập nhật. Vui lòng kiểm tra IDPhieu/MaTC.")

    if pending_updates:
        worksheet.batch_update(pending_updates, value_input_option="USER_ENTERED")

    saved_details = list(current_by_key.values())
    total = calculate_total_score(saved_details)
    update_header_fields(
        id_phieu,
        {
            "TongDiem": total,
            "NgayCapNhat": timestamp,
            "NguoiCapNhatCuoi": ma_gv,
        },
    )
    return total

def validate_before_submit(details: list[dict], criteria: list[dict] | None = None) -> list[str]:
    """Return teacher-facing warning messages before the form is submitted.

    The warning uses MaHienThi when available so teachers can locate the row
    shown on the score form. This function only warns; submit_phieu() does not
    block submission on these warnings.
    """
    display_by_ma_tc = _build_display_code_map(criteria or [])
    errors: list[str] = []

    for detail in details:
        record = _canonicalize_ct_record(detail)
        diem_mac_dinh = _safe_float(record.get("DiemMacDinh"))
        diem_gv = _safe_float(record.get("DiemGV"))
        ghi_chu_gv = str(record.get("GhiChuGV", "")).strip()
        ma_tc = str(record.get("MaTC", "")).strip()

        if diem_gv != diem_mac_dinh and not ghi_chu_gv:
            display_code = display_by_ma_tc.get(ma_tc) or ma_tc
            errors.append(f"Tiêu chí {display_code}: cần nhập giải trình.")
    return errors





def _build_display_code_map(criteria: list[dict]) -> dict[str, str]:
    """Map MaTC to teacher-facing display code, e.g. parent 2.1 + child b -> 2.1b."""
    nodes: dict[str, dict[str, str]] = {}
    for criterion in criteria:
        ma_tc = _get_text(
            criterion,
            "ma_tc",
            "MaTC",
            "Mã TC",
            "MaTieuChi",
            "Mã tiêu chí",
            "ma_tieu_chi",
        )
        if not ma_tc:
            continue
        nodes[ma_tc] = {
            "parent": _get_text(criterion, "ma_cha", "MaCha", "Mã cha", "ma_cha_tieu_chi"),
            "display": _get_text(criterion, "ma_hien_thi", "MaHienThi", "Mã hiển thị", "display_code"),
        }

    resolved: dict[str, str] = {}

    def resolve(ma_tc: str, visiting: set[str] | None = None) -> str:
        if ma_tc in resolved:
            return resolved[ma_tc]
        visiting = visiting or set()
        if ma_tc in visiting:
            return ma_tc
        visiting.add(ma_tc)

        node = nodes.get(ma_tc, {})
        display = str(node.get("display", "")).strip()
        parent_code = str(node.get("parent", "")).strip()
        parent_display = resolve(parent_code, visiting) if parent_code in nodes else ""

        if not display:
            value = parent_display or ma_tc
        elif not parent_display:
            value = display
        elif display.startswith(parent_display):
            value = display
        elif _looks_like_full_display_code(display):
            value = display
        elif len(display) <= 3 and display.replace(".", "").isalnum():
            value = f"{parent_display}{display}"
        else:
            value = display

        resolved[ma_tc] = value
        return value

    for ma_tc in nodes:
        resolve(ma_tc)

    return resolved


def _looks_like_full_display_code(value: str) -> bool:
    text = str(value).strip()
    return "." in text or text.upper() in {"I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"}

def _trace_submit_event(trace: list[str] | None, message: str) -> None:
    """Append one submit-flow trace line for temporary debugging."""
    if trace is not None:
        trace.append(str(message))


def _format_submit_trace(trace: list[str]) -> str:
    return "\n".join(trace)

def submit_phieu(id_phieu: str, ma_gv: str, changed_rows: list[dict] | None = None) -> dict:
    """Save final changes, lock teacher editing, and audit submit action."""
    saved_total: float | None = None
    if changed_rows:
        saved_total = save_teacher_scores(id_phieu, changed_rows, ma_gv)
        details = changed_rows
    else:
        details = get_chi_tiet_phieu(id_phieu)

    clear_sheet_records_cache(SHEET_TH)
    phieu = get_header_by_id(id_phieu)
    if not phieu:
        raise ValueError("Không tìm thấy phiếu thi đua cần nộp.")

    current_status = _safe_int(phieu.get("TrangThai"), default=0)
    current_locked = _parse_bool(phieu.get("KhoaGV"))
    if current_locked or current_status == STATUS_DA_NOP:
        # The form is already submitted/locked. Return the current header instead
        # of raising an error so the page can refresh into read-only mode.
        if current_status != STATUS_DA_NOP:
            update_header_fields(id_phieu, {"TrangThai": STATUS_DA_NOP})
            clear_sheet_records_cache(SHEET_TH)
            repaired = get_header_by_id(id_phieu)
            return repaired or phieu
        return phieu

    timestamp = now_iso()
    lan_nop = int(str(phieu.get("LanNop") or 0).strip() or 0) + 1
    total = saved_total if saved_total is not None else calculate_total_score(details)
    update_header_fields(
        id_phieu,
        {
            "TrangThai": STATUS_DA_NOP,
            "TongDiem": total,
            "KhoaGV": True,
            "NgayCapNhat": timestamp,
            "NgayNop": timestamp,
            "NguoiCapNhatCuoi": ma_gv,
            "NguoiNop": ma_gv,
            "LanNop": lan_nop,
        },
    )
    clear_sheet_records_cache(SHEET_TH)
    refreshed = get_header_by_id(id_phieu)
    if not refreshed:
        raise ValueError("Không đọc lại được TH_ThiDua sau khi cập nhật.")

    refreshed_status = int(str(refreshed.get("TrangThai") or 0).strip() or 0)
    refreshed_locked = _parse_bool(refreshed.get("KhoaGV"))
    if refreshed_status != STATUS_DA_NOP or not refreshed_locked:
        raise ValueError(
            "TH_ThiDua chưa chuyển sang Đã nộp/Khóa GV sau khi cập nhật. "
            f"TrangThai={refreshed.get('TrangThai')} | KhoaGV={refreshed.get('KhoaGV')}"
        )

    write_audit_log(ACTION_SUBMIT_PHIEU, id_phieu, ma_gv, "Giáo viên xác nhận nộp phiếu thi đua.")
    clear_sheet_records_cache(SHEET_TH)
    return refreshed

def is_teacher_allowed_to_create() -> bool:
    """Return True if teacher can self-create a form today."""
    return is_scoring_window_open()


def is_teacher_allowed_to_edit(phieu: dict) -> bool:
    """Return True if teacher can edit the current form."""
    if not phieu:
        return False
    if _parse_bool(phieu.get("KhoaGV")):
        return False
    if int(phieu.get("TrangThai") or 0) == STATUS_DA_CHOT:
        return False
    return True


def get_header_by_id(id_phieu: str) -> dict | None:
    """Return one TH_ThiDua row by ID, normalized to canonical keys."""
    for record in read_sheet_records(SHEET_TH):
        canonical = _canonicalize_th_record(record)
        if str(canonical.get("ID", "")).strip() == str(id_phieu).strip():
            return canonical
    return None

def update_header_fields(id_phieu: str, fields: dict[str, Any]) -> None:
    """Update selected TH_ThiDua fields by canonical header names using batch_update."""
    worksheet = get_worksheet(SHEET_TH)
    records = read_sheet_records(SHEET_TH)
    header_map = _get_header_map(worksheet)

    row_number = None
    for index, record in enumerate(records):
        if _get_th_record_text(record, "ID") == str(id_phieu).strip():
            row_number = index + 2
            break

    if row_number is None:
        raise ValueError(f"Không tìm thấy TH_ThiDua.ID = {id_phieu}")

    pending_updates: list[dict[str, Any]] = []
    missing_required: list[str] = []
    for field, value in fields.items():
        column_number = header_map.get(field)
        if column_number:
            pending_updates.append({"range": _a1(row_number, column_number), "values": [[value]]})
        elif field in {"TrangThai", "KhoaGV", "TongDiem", "NgayNop"}:
            missing_required.append(field)

    if missing_required:
        available = ", ".join(sorted(str(key) for key in header_map.keys()))
        raise ValueError(
            "TH_ThiDua thiếu cột bắt buộc để cập nhật phiếu: "
            + ", ".join(missing_required)
            + f". Các cột đọc được: {available}"
        )

    if pending_updates:
        worksheet.batch_update(pending_updates, value_input_option="USER_ENTERED")
        clear_sheet_records_cache(SHEET_TH)

def _header_by_column(header_map: dict[str, int], column_number: int) -> str:
    for header, number in header_map.items():
        if number == column_number:
            return header
    return ""


def _canonical_header_name(header: str) -> str:
    normalized = _normalize_key(header)
    aliases = {
        # Common / TH_ThiDua
        "id": "ID",
        "nam_hoc": "NamHoc",
        "namhoc": "NamHoc",
        "thang": "Thang",
        "ma_gv": "MaGV",
        "magv": "MaGV",
        "trang_thai": "TrangThai",
        "trangthai": "TrangThai",
        "trang_thai_phieu": "TrangThai",
        "trangthaiphieu": "TrangThai",
        "tong_diem": "TongDiem",
        "tongdiem": "TongDiem",
        "xep_loai": "XepLoai",
        "xeploai": "XepLoai",
        "ngay_tao": "NgayTao",
        "ngaytao": "NgayTao",
        "ngay_nop": "NgayNop",
        "ngaynop": "NgayNop",
        "nguoi_cap_nhat": "NguoiCapNhatCuoi",
        "nguoicapnhat": "NguoiCapNhatCuoi",
        "nguoi_cap_nhat_cuoi": "NguoiCapNhatCuoi",
        "nguoicapnhatcuoi": "NguoiCapNhatCuoi",
        "thoi_gian_cap_nhat": "NgayCapNhat",
        "thoigiancapnhat": "NgayCapNhat",
        "ngay_cap_nhat": "NgayCapNhat",
        "ngaycapnhat": "NgayCapNhat",
        "ghi_chu": "GhiChuBGH",
        "ghichu": "GhiChuBGH",
        "khoa_gv": "KhoaGV",
        "khoagv": "KhoaGV",
        "khoa_bgh": "KhoaBGH",
        "khoabgh": "KhoaBGH",
        "nguoi_nop": "NguoiNop",
        "nguoinop": "NguoiNop",
        "lan_nop": "LanNop",
        "lannop": "LanNop",
        "lan_mo_khoa": "LanMoKhoa",
        "lanmokhoa": "LanMoKhoa",
        # CT_ThiDua
        "idphieu": "IDPhieu",
        "id_phieu": "IDPhieu",
        "ma_tc": "MaTC",
        "matc": "MaTC",
        "ma_tieu_chi": "MaTC",
        "diem_mac_dinh": "DiemMacDinh",
        "diemmacdinh": "DiemMacDinh",
        "diem_goi_y": "DiemMacDinh",
        "diem_gv": "DiemGV",
        "diemgv": "DiemGV",
        "diem_giao_vien": "DiemGV",
        "tu_cham": "DiemGV",
        "diem_bgh": "DiemBGH",
        "diembgh": "DiemBGH",
        "diem_ban_giam_hieu": "DiemBGH",
        "ghi_chu_gv": "GhiChuGV",
        "ghichugv": "GhiChuGV",
        "ghi_chu_giao_vien": "GhiChuGV",
        "ghi_chu_bgh": "GhiChuBGH",
        "ghichubgh": "GhiChuBGH",
        "ghi_chu_ban_giam_hieu": "GhiChuBGH",
    }
    compact = normalized.replace("_", "")
    if normalized in aliases:
        return aliases[normalized]
    if compact in aliases:
        return aliases[compact]
    for expected in TH_HEADERS + CT_HEADERS:
        if normalized == _normalize_key(expected) or compact == _normalize_key(expected).replace("_", ""):
            return expected
    return str(header).strip()

def _a1(row_number: int, column_number: int) -> str:
    """Convert 1-based row/column numbers to A1 notation."""
    letters = ""
    col = int(column_number)
    while col:
        col, remainder = divmod(col - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{int(row_number)}"


def _get_header_map(worksheet: Any) -> dict[str, int]:
    headers = worksheet.row_values(1)
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        raw_header = str(header).strip()
        if not raw_header:
            continue
        canonical = _canonical_header_name(raw_header)
        # Prefer canonical names used by the code while still allowing actual
        # Google Sheets headers to vary slightly in spelling/spacing.
        header_map.setdefault(canonical, index + 1)
        header_map.setdefault(raw_header, index + 1)
    return header_map


def _update_cells_by_headers(worksheet: Any, row_number: int, header_map: dict[str, int], fields: dict[str, Any]) -> None:
    pending_updates: list[dict[str, Any]] = []
    for field, value in fields.items():
        column_number = header_map.get(field)
        if column_number:
            pending_updates.append({"range": _a1(row_number, column_number), "values": [[value]]})
    if pending_updates:
        worksheet.batch_update(pending_updates, value_input_option="USER_ENTERED")

def _get_text(record: dict, *keys: str) -> str:
    normalized = {_normalize_key(key): value for key, value in record.items()}
    for key in keys:
        value = record.get(key)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
        normalized_value = normalized.get(_normalize_key(key))
        if normalized_value is not None and str(normalized_value).strip() != "":
            return str(normalized_value).strip()
    return ""


def _normalize_month_key(value: Any) -> str:
    """Normalize month values to MM/YYYY when possible for safe comparison."""
    text = str(value or "").strip().replace("'", "")
    if not text:
        return ""

    parts = text.split("/")
    if len(parts) == 2:
        month, year = parts[0].strip(), parts[1].strip()
        try:
            return f"{int(float(month)):02d}/{int(float(year)):04d}"
        except ValueError:
            return text

    compact = "".join(ch for ch in text if ch.isdigit())
    if len(compact) == 6:
        return f"{compact[:2]}/{compact[2:]}"

    return text


def _normalize_key(value: str) -> str:
    text = str(value).strip().lower()
    replacements = {
        "ã": "a", "à": "a", "á": "a", "ả": "a", "ạ": "a", "ă": "a", "ằ": "a", "ắ": "a", "ẳ": "a", "ặ": "a", "â": "a", "ầ": "a", "ấ": "a", "ẩ": "a", "ậ": "a",
        "đ": "d", "è": "e", "é": "e", "ẻ": "e", "ẹ": "e", "ê": "e", "ề": "e", "ế": "e", "ể": "e", "ệ": "e",
        "ì": "i", "í": "i", "ỉ": "i", "ị": "i", "ò": "o", "ó": "o", "ỏ": "o", "ọ": "o", "ô": "o", "ồ": "o", "ố": "o", "ổ": "o", "ộ": "o", "ơ": "o", "ờ": "o", "ớ": "o", "ở": "o", "ợ": "o",
        "ù": "u", "ú": "u", "ủ": "u", "ụ": "u", "ư": "u", "ừ": "u", "ứ": "u", "ử": "u", "ự": "u", "ỳ": "y", "ý": "y", "ỷ": "y", "ỵ": "y",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return "_".join(text.replace("(", " ").replace(")", " ").replace("/", " ").split())


def _get_float(record: dict, *keys: str) -> float:
    text = _get_text(record, *keys)
    if not text:
        return 0.0
    return float(str(text).replace(",", "."))



def _safe_float(value: Any) -> float:
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


_TH_ALIASES = {
    "ID": ("ID", "id"),
    "NamHoc": ("NamHoc", "Năm học", "nam_hoc", "namhoc"),
    "Thang": ("Thang", "Tháng", "thang"),
    "MaGV": ("MaGV", "Mã GV", "ma_gv", "magv"),
    "TrangThai": ("TrangThai", "Trạng thái phiếu", "Trạng thái", "trang_thai", "trangthai", "trang_thai_phieu"),
    "TongDiem": ("TongDiem", "Tổng điểm", "tong_diem", "tongdiem"),
    "KhoaGV": ("KhoaGV", "Khóa GV", "khoa_gv", "khoagv"),
    "KhoaBGH": ("KhoaBGH", "Khóa BGH", "khoa_bgh", "khoabgh"),
    "NgayTao": ("NgayTao", "Ngày tạo", "ngay_tao", "ngaytao"),
    "NgayCapNhat": ("NgayCapNhat", "Thời gian cập nhật", "Ngày cập nhật", "ngay_cap_nhat", "ngaycapnhat"),
    "NgayNop": ("NgayNop", "Ngày nộp", "ngay_nop", "ngaynop"),
    "NguoiCapNhatCuoi": ("NguoiCapNhatCuoi", "Người cập nhật", "Người cập nhật cuối", "nguoi_cap_nhat", "nguoi_cap_nhat_cuoi"),
    "NguoiNop": ("NguoiNop", "Người nộp", "nguoi_nop", "nguoinop"),
    "GhiChuBGH": ("GhiChuBGH", "Ghi chú", "Ghi chú BGH", "ghi_chu_bgh", "ghichubgh"),
    "LanNop": ("LanNop", "Lần nộp", "lan_nop", "lannop"),
    "LanMoKhoa": ("LanMoKhoa", "Lần mở khóa", "lan_mo_khoa", "lanmokhoa"),
}


def _canonicalize_th_record(record: dict) -> dict:
    """Return TH_ThiDua row with canonical field names.

    Some legacy rows can be partially updated: KhoaGV/NgayNop may already
    indicate a submitted form while TrangThai still contains 2.  For app logic,
    a locked teacher form with a submit timestamp must be treated as submitted
    so the page becomes read-only and displays the correct status.
    """
    canonical = {name: _get_text(record, *aliases) for name, aliases in _TH_ALIASES.items()}

    status = _safe_int(canonical.get("TrangThai"), default=0)
    locked = _parse_bool(canonical.get("KhoaGV"))
    ngay_nop = str(canonical.get("NgayNop", "")).strip()

    if status != STATUS_DA_CHOT and status != STATUS_DA_NOP and (locked or ngay_nop):
        canonical["TrangThai"] = str(STATUS_DA_NOP)

    return canonical


def _safe_int(value: Any, default: int = 0) -> int:
    text = str(value or "").strip()
    if not text:
        return default
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return default


def _get_th_record_text(record: dict, canonical: str) -> str:
    aliases = _TH_ALIASES.get(canonical, (canonical,))
    return _get_text(record, *aliases)


_CT_ALIASES = {
    "ID": ("ID", "id"),
    "IDPhieu": ("IDPhieu", "ID Phiếu", "id_phieu", "idphieu"),
    "MaTC": ("MaTC", "Mã TC", "ma_tc", "MaTieuChi", "Mã tiêu chí", "ma_tieu_chi"),
    "DiemMacDinh": ("DiemMacDinh", "Điểm mặc định", "Điểm gợi ý", "diem_mac_dinh", "diemmacdinh"),
    "DiemGV": ("DiemGV", "Điểm GV", "Điểm giáo viên", "Điểm GV chấm", "Tự chấm", "diem_gv", "diemgv"),
    "DiemBGH": ("DiemBGH", "Điểm BGH", "Điểm BGH chỉnh", "diem_bgh", "diembgh"),
    "GhiChuGV": ("GhiChuGV", "Ghi chú GV", "Ghi chú", "ghi_chu_gv", "ghichugv"),
    "GhiChuBGH": ("GhiChuBGH", "Ghi chú BGH", "ghi_chu_bgh", "ghichubgh"),
    "NgayCapNhat": ("NgayCapNhat", "Ngày cập nhật", "ngay_cap_nhat"),
    "NguoiCapNhat": ("NguoiCapNhat", "Người cập nhật", "nguoi_cap_nhat"),
}


def _get_canonical_header_map(worksheet: Any) -> dict[str, int]:
    headers = worksheet.row_values(1)
    normalized_headers = {_normalize_key(header): index + 1 for index, header in enumerate(headers)}
    result: dict[str, int] = {}
    for canonical, aliases in _CT_ALIASES.items():
        for alias in aliases:
            column_number = normalized_headers.get(_normalize_key(alias))
            if column_number:
                result[canonical] = column_number
                break
    return result


def _canonicalize_ct_record(record: dict) -> dict:
    return {canonical: _get_text(record, *aliases) for canonical, aliases in _CT_ALIASES.items()}


def _get_record_text(record: dict, canonical: str) -> str:
    aliases = _CT_ALIASES.get(canonical, (canonical,))
    return _get_text(record, *aliases)

def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y", "x", "co", "có"}
